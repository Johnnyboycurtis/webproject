from django.shortcuts import render, redirect
from mainapp.forms import DataForm
from mainapp import models

def index(request):
    form = DataForm()
    queryset = models.Data.objects.all()
    if request.method == 'POST':
        form = DataForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    context = {'title': 'Simple App', 'form': form, 'posts': queryset}
    return render(request, 'mainapp/index.html', context=context)


from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font

def export_data(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

    # create workbook
    wb = Workbook()
    sheet = wb.active

    # stylize header row
    # 'id','title', 'quantity','pub_date'
      
    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "id"
    c1.font = Font(bold=True)
    
    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "title"
    c2.font = Font(bold=True)

    c3 = sheet.cell(row= 1 , column = 3) 
    c3.value = "quantity"
    c3.font = Font(bold=True)

    c4 = sheet.cell(row= 1 , column = 4) 
    c4.value = "pub_date"
    c4.font = Font(bold=True)
    
    # export data to Excel
    rows = models.Data.objects.all().values_list('id','category', 'quantity','pub_date',)
    for row_num, row in enumerate(rows, 1):
        # row is just a tuple
        for col_num, value in enumerate(row):
            c5 = sheet.cell(row=row_num+1, column=col_num+1) 
            c5.value = value

    wb.save(response)

    return response